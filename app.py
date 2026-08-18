import io
import os
import secrets
from datetime import datetime, timedelta

import openpyxl
import qrcode
from dotenv import load_dotenv
from flask import Flask, render_template, request, redirect, url_for, abort, jsonify, send_file, session
from PIL import Image, UnidentifiedImageError

import storage
import theme
import wallet
import wallet_apple
import sso_config
import sso_photo_cache
from card_render import render_card_front, render_card_back
from admin import admin_bp
from sso_routes import sso_bp

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or "dev-only-change-me"
# Explicit cookie attributes (rather than Flask's defaults) so the OAuth
# "state" cookie set by /auth/microsoft/login reliably survives the
# redirect round-trip to Microsoft and back -- a dropped state cookie is
# what surfaces as "Sign-in session expired or is invalid".
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = not app.debug
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)
app.register_blueprint(admin_bp)
app.register_blueprint(sso_bp)


@app.context_processor
def inject_theme():
    is_preview = bool(session.get("is_admin")) and request.args.get("preview_theme") == "1"
    active = theme.active_theme(is_preview)
    nav_flags = {item["id"]: item.get("enabled", True) for item in active.get("nav_links", [])}
    reg_fields = {item["id"]: item for item in active.get("registration_fields", [])}
    return {
        "site_theme": active,
        "site_theme_css": theme.css_overrides(active),
        "site_theme_preview": is_preview,
        "site_theme_nav": nav_flags,
        "site_theme_reg_fields": reg_fields,
    }

PORTAL_BASE_URL = (os.environ.get("PORTAL_BASE_URL") or "http://127.0.0.1:5000").rstrip("/")

# All generated data (Excel + QR codes + photos) lives outside the project
# folder so it survives redeploys and can be pointed at a shared location later.
# Vercel's filesystem is read-only except /tmp, so always use that there (it
# sets VERCEL=1) regardless of DATA_ROOT — note /tmp is wiped between
# invocations, so the Excel "database" and uploaded photos will NOT persist
# on Vercel. Elsewhere, honor DATA_ROOT only if it's actually set to
# something (os.environ.get's default doesn't apply to a blank env var).
if os.environ.get("VERCEL"):
    DATA_ROOT = "/tmp"
else:
    DATA_ROOT = os.environ.get("DATA_ROOT") or r"C:\MDX-Digital-ID\Test"
QR_DIR = os.path.join(DATA_ROOT, "qrcodes")
PHOTO_DIR = os.path.join(DATA_ROOT, "photos")
EXCEL_PATH = os.path.join(DATA_ROOT, "Staff_Digital_ID.xlsx")

os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(PHOTO_DIR, exist_ok=True)

MAX_PHOTO_BYTES = 5 * 1024 * 1024
ALLOWED_PHOTO_FORMATS = {"JPEG", "PNG", "WEBP"}
PHOTO_MAX_DIMENSION = 640

DEPARTMENTS = [
    "Visa",
    "Human Resource Management",
    "Psychology",
    "Accounting & Finance",
    "Business School",
    "Computer Engineering & Informatics",
    "Finance",
    "Facilities",
    "Quality",
    "International Foundation Programme",
    "Directors",
    "Campus Central",
    "Centre for Academic Success",
    "Student Activities",
    "Library",
    "Media",
    "Marketing",
    "Law",
    "Graphic Design",
    "CIPD",
    "School of Health and Education",
    "Dubai Academic Registry",
    "Careers & Employability Services",
    "Student Recruitment",
    "Information Technology, AI and Cybersecurity (ITAC)",
    "Digital Marketing",
    "Digital Transformation",
    "London Sports Institute",
    "HR",
    "Tourism",
]
EMPLOYMENT_STATUSES = ["Full-Time", "Part-Time", "Contract", "Visiting"]
GENDERS = ["Male", "Female"]
CATEGORIES = ["Administration", "Faculty"]

DIRECTORY_COLUMNS = [
    "Staff ID", "First Name", "Last Name", "Email",
    "Department", "Job Title", "Gender", "Employment Status",
]
DIGITAL_ID_COLUMNS = [
    "Token", "Track ID", "Staff ID", "First Name", "Last Name", "Email", "Mobile Number",
    "Department", "Job Title", "Gender", "Employment Status",
    "Photo Filename", "QR Filename", "Created At", "Status", "Microsoft User ID",
    "Category", "UK IT User ID", "Local Login", "MISIS",
]
DIGITAL_ID_STATUSES = ("Active", "Suspended", "Deactivated", "Expired")
ACTIVITY_LOG_COLUMNS = ["Event", "Actor", "Detail", "Created At"]

SAMPLE_DIRECTORY_ROWS = [
    ["MDX00001", "Sample", "Employee", "sample.employee@mdx.ac.ae",
     "Information Technology, AI and Cybersecurity (ITAC)", "IT Support Engineer", "Male", "Full-Time"],
    ["MDX00002", "Sample", "Faculty", "sample.faculty@mdx.ac.ae",
     "Academic - Business School", "Senior Lecturer", "Female", "Full-Time"],
]


def _sheet_has_data_rows(ws):
    return ws.max_row > 1


def init_excel():
    """Creates/migrates the workbook. Only resets a sheet's header if that
    sheet has no data rows yet, so real staff data is never wiped on restart."""
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws_dir = wb.active
        ws_dir.title = "Staff_Directory"
        ws_dir.append(DIRECTORY_COLUMNS)
        for row in SAMPLE_DIRECTORY_ROWS:
            ws_dir.append(row)
        ws_ids = wb.create_sheet("Digital_IDs")
        ws_ids.append(DIGITAL_ID_COLUMNS)
        ws_log = wb.create_sheet("Activity_Log")
        ws_log.append(ACTIVITY_LOG_COLUMNS)
        wb.save(EXCEL_PATH)
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    changed = False

    if "Staff_Directory" not in wb.sheetnames:
        ws = wb.create_sheet("Staff_Directory")
        ws.append(DIRECTORY_COLUMNS)
        for row in SAMPLE_DIRECTORY_ROWS:
            ws.append(row)
        changed = True
    else:
        ws = wb["Staff_Directory"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if header != DIRECTORY_COLUMNS and not _sheet_has_data_rows(ws):
            wb.remove(ws)
            ws = wb.create_sheet("Staff_Directory")
            ws.append(DIRECTORY_COLUMNS)
            for row in SAMPLE_DIRECTORY_ROWS:
                ws.append(row)
            changed = True

    if "Digital_IDs" not in wb.sheetnames:
        ws = wb.create_sheet("Digital_IDs")
        ws.append(DIGITAL_ID_COLUMNS)
        changed = True
    else:
        ws = wb["Digital_IDs"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if header != DIGITAL_ID_COLUMNS:
            if not _sheet_has_data_rows(ws):
                wb.remove(ws)
                ws = wb.create_sheet("Digital_IDs")
                ws.append(DIGITAL_ID_COLUMNS)
                changed = True
            elif header == DIGITAL_ID_COLUMNS[: len(header)]:
                # Existing rows are untouched; new columns just get appended
                # header cells so future rows line up under a real name too.
                for i, col in enumerate(DIGITAL_ID_COLUMNS[len(header):], start=len(header) + 1):
                    ws.cell(row=1, column=i, value=col)
                changed = True

    if "Activity_Log" not in wb.sheetnames:
        ws = wb.create_sheet("Activity_Log")
        ws.append(ACTIVITY_LOG_COLUMNS)
        changed = True

    if "Sheet" in wb.sheetnames and not _sheet_has_data_rows(wb["Sheet"]) and wb["Sheet"]["A1"].value is None:
        wb.remove(wb["Sheet"])
        changed = True

    if changed:
        wb.save(EXCEL_PATH)


# Postgres + Vercel Blob when a database is configured (required on Vercel,
# since /tmp isn't shared or persistent across invocations); otherwise the
# original Excel + local-file storage below (used for the LAN deployment).
if storage.db_configured():
    storage.init_db(SAMPLE_DIRECTORY_ROWS)
else:
    init_excel()


def _rows_as_dicts(ws, columns):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        yield dict(zip(columns, row))


def find_directory_record_by_staff_id(staff_id):
    if storage.db_configured():
        return storage.find_directory_record_by_staff_id(staff_id)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Staff_Directory"]
    staff_id = (staff_id or "").strip().lower()
    for record in _rows_as_dicts(ws, DIRECTORY_COLUMNS):
        if str(record["Staff ID"]).strip().lower() == staff_id:
            return record
    return None


def find_active_digital_id_by_staff(staff_id):
    if storage.db_configured():
        return storage.find_active_digital_id_by_staff(staff_id)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    staff_id = (staff_id or "").strip().lower()
    for record in _rows_as_dicts(ws, DIGITAL_ID_COLUMNS):
        if str(record["Staff ID"]).strip().lower() == staff_id and record["Status"] == "Active":
            return record
    return None


def find_digital_id_by_token(token):
    if storage.db_configured():
        return storage.find_digital_id_by_token(token)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    for record in _rows_as_dicts(ws, DIGITAL_ID_COLUMNS):
        if record["Token"] == token:
            return record
    return None


def find_digital_id_by_track(track_id):
    track_id = (track_id or "").strip().lower()
    if not track_id:
        return None
    if storage.db_configured():
        return storage.find_digital_id_by_track(track_id)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    for record in _rows_as_dicts(ws, DIGITAL_ID_COLUMNS):
        if str(record.get("Track ID") or "").strip().lower() == track_id:
            return record
    return None


def find_digital_id_by_ms_user_id(ms_user_id):
    if not ms_user_id:
        return None
    if storage.db_configured():
        return storage.find_digital_id_by_ms_user_id(ms_user_id)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    for record in _rows_as_dicts(ws, DIGITAL_ID_COLUMNS):
        if str(record.get("Microsoft User ID") or "").strip() == str(ms_user_id).strip():
            return record
    return None


def list_digital_ids(search=None):
    if storage.db_configured():
        return storage.list_digital_ids(search)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    records = list(_rows_as_dicts(ws, DIGITAL_ID_COLUMNS))
    records.sort(key=lambda r: r.get("Created At") or "", reverse=True)
    if search:
        needle = search.strip().lower()
        records = [
            r for r in records
            if needle in str(r.get("Staff ID", "")).lower()
            or needle in f'{r.get("First Name", "")} {r.get("Last Name", "")}'.lower()
            or needle in str(r.get("Email", "")).lower()
        ]
    return records


def count_active_digital_ids():
    return sum(1 for r in list_digital_ids() if r.get("Status") == "Active")


def update_digital_id_status(token, status):
    if status not in DIGITAL_ID_STATUSES:
        return
    if storage.db_configured():
        storage.update_digital_id_status(token, status)
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    token_idx = DIGITAL_ID_COLUMNS.index("Token")
    status_idx = DIGITAL_ID_COLUMNS.index("Status")
    for row in ws.iter_rows(min_row=2):
        if row[token_idx].value == token:
            row[status_idx].value = status
            break
    wb.save(EXCEL_PATH)


def log_event(event_type, actor, detail):
    """Append-only activity trail: admin logins, staff Microsoft sign-ins,
    and Digital ID create/update events, surfaced on /admin/activity."""
    created_at = datetime.now().isoformat(timespec="seconds")
    if storage.db_configured():
        storage.log_event(event_type, actor or "", detail or "", created_at)
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Activity_Log"]
    ws.append([event_type, actor or "", detail or "", created_at])
    wb.save(EXCEL_PATH)


def list_activity_log(limit=200):
    if storage.db_configured():
        return storage.list_activity_log(limit)
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Activity_Log"]
    rows = list(_rows_as_dicts(ws, ACTIVITY_LOG_COLUMNS))
    rows.sort(key=lambda r: r.get("Created At") or "", reverse=True)
    return rows[:limit]


def update_digital_id_record(token, updates):
    """updates: dict of DIGITAL_ID_COLUMNS display names -> new values."""
    if storage.db_configured():
        sql_by_display = dict(zip(storage.DIGITAL_ID_DISPLAY, storage.DIGITAL_ID_COLUMNS_SQL))
        storage.update_digital_id_fields(token, {sql_by_display[k]: v for k, v in updates.items()})
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    token_idx = DIGITAL_ID_COLUMNS.index("Token")
    for row in ws.iter_rows(min_row=2):
        if row[token_idx].value == token:
            for key, value in updates.items():
                row[DIGITAL_ID_COLUMNS.index(key)].value = value
            break
    wb.save(EXCEL_PATH)


def regenerate_digital_id_qr(token):
    """Rebuilds the vCard QR for an existing Digital ID from its current
    record fields -- used by the admin 'Regenerate QR Code' action, e.g.
    after a staff member's job title or department changes."""
    record = find_digital_id_by_token(token)
    if not record:
        return

    vcard_payload = build_vcard(
        first_name=record["First Name"], last_name=record["Last Name"],
        job_title=record["Job Title"], department=record["Department"],
        email=record["Email"], mobile=record.get("Mobile Number"),
    )
    qr = qrcode.QRCode(version=1, box_size=10, border=3)
    qr.add_data(vcard_payload)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#000000", back_color="#ffffff")
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_bytes = qr_buf.getvalue()

    if storage.blob_configured():
        qr_value = storage.upload_qr(qr_bytes, token)
    else:
        qr_filename = f"qr_{token}.png"
        with open(os.path.join(QR_DIR, qr_filename), "wb") as f:
            f.write(qr_bytes)
        qr_value = qr_filename

    if storage.db_configured():
        conn = storage.get_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE digital_ids SET qr_url = %s WHERE token = %s", (qr_value, token))
            conn.commit()
        finally:
            conn.close()
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    token_idx = DIGITAL_ID_COLUMNS.index("Token")
    qr_idx = DIGITAL_ID_COLUMNS.index("QR Filename")
    for row in ws.iter_rows(min_row=2):
        if row[token_idx].value == token:
            row[qr_idx].value = qr_value
            break
    wb.save(EXCEL_PATH)


def generate_track_id():
    """Short, memorable reference code shown to the staff member — distinct
    from the long secure Token used inside the QR/verify URL."""
    while True:
        candidate = "MDX-" + secrets.token_hex(3).upper()
        if not find_digital_id_by_track(candidate):
            return candidate


def build_vcard(first_name, last_name, job_title, department, email, mobile=None):
    """vCard 3.0 payload for the staff-contact QR — a real, scannable
    business card, not a security/verification link."""
    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"FN:{first_name} {last_name}".strip(),
        "ORG:Middlesex University Dubai",
    ]
    if job_title:
        lines.append(f"TITLE:{job_title}")
    if email:
        lines.append(f"EMAIL;TYPE=WORK:{email}")
    if mobile:
        lines.append(f"TEL;TYPE=CELL:{mobile}")
    lines.append("URL:https://www.mdx.ac.ae")
    if department:
        lines.append(f"NOTE:Department - {department}")
    lines.append("END:VCARD")
    return "\r\n".join(lines)


def append_digital_id(record):
    if storage.db_configured():
        storage.append_digital_id(record)
        return
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Digital_IDs"]
    ws.append([record[col] for col in DIGITAL_ID_COLUMNS])
    wb.save(EXCEL_PATH)


def _load_media_bytes(filename_or_url, local_dir):
    """Photo/QR Filename holds either a bare local filename (Excel/local-file
    mode) or a full Vercel Blob URL (Postgres/Blob mode) -- fetch whichever
    it is into an in-memory file-like object."""
    if not filename_or_url:
        return None
    if filename_or_url.startswith("http://") or filename_or_url.startswith("https://"):
        return io.BytesIO(storage.download_blob(filename_or_url))
    path = os.path.join(local_dir, filename_or_url)
    return open(path, "rb") if os.path.exists(path) else None


REQUIRED_FIELDS = [
    "first_name", "last_name", "staff_id", "email",
    "department", "job_title", "gender", "employment_status",
]


def _validate_and_save_photo(file_storage, token):
    """Validates format/size, downscales, and saves as a JPEG.
    Returns (filename, error_message) \u2014 exactly one will be set."""
    if not file_storage or not file_storage.filename:
        return None, "Staff photo is required."

    data = file_storage.read()
    if len(data) > MAX_PHOTO_BYTES:
        return None, "Photo is too large (max 5MB)."

    try:
        image = Image.open(io.BytesIO(data))
        image.verify()
        image = Image.open(io.BytesIO(data))  # verify() consumes the parser; reopen to actually use it
    except UnidentifiedImageError:
        return None, "Unsupported file \u2014 please upload a JPEG, PNG, or WebP image."

    if image.format not in ALLOWED_PHOTO_FORMATS:
        return None, "Unsupported format \u2014 please upload a JPEG, PNG, or WebP image."

    image = image.convert("RGB")
    image.thumbnail((PHOTO_MAX_DIMENSION, PHOTO_MAX_DIMENSION))
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=88)
    photo_bytes = buf.getvalue()

    if storage.blob_configured():
        return storage.upload_photo(photo_bytes, token), None

    filename = f"{token}.jpg"
    with open(os.path.join(PHOTO_DIR, filename), "wb") as f:
        f.write(photo_bytes)
    return filename, None


def _handle_submission():
    form = request.form
    values = {field: (form.get(field) or "").strip() for field in REQUIRED_FIELDS}
    mobile_number = (form.get("mobile_number") or "").strip()
    category = (form.get("category") or "").strip()
    uk_it_user_id = (form.get("uk_it_user_id") or "").strip()
    local_login = (form.get("local_login") or "").strip()
    misis = (form.get("misis") or "").strip()

    edit_token = (form.get("edit_token") or "").strip()
    edit_record = None
    if edit_token:
        edit_record = find_digital_id_by_token(edit_token)
        ms_session_id = session.get("ms_user_id")
        if not edit_record or not ms_session_id or edit_record.get("Microsoft User ID") != ms_session_id:
            abort(403)

    missing = [f for f in REQUIRED_FIELDS if not values[f]]
    photo_file = request.files.get("photo")
    photo_provided = bool(photo_file and photo_file.filename)
    # A photo is required to create a new Digital ID, but editing an
    # existing one can keep the photo already on file.
    if not photo_provided and not edit_record:
        missing.append("photo")

    if missing:
        return render_template(
            "index.html",
            departments=DEPARTMENTS, employment_statuses=EMPLOYMENT_STATUSES, genders=GENDERS, categories=CATEGORIES,
            error=f"Missing required field(s): {', '.join(missing)}",
            form=form, edit_record=edit_record,
        ), 400

    ms_user_id = (session.get("ms_user_id") or "").strip()

    if not edit_record and (find_active_digital_id_by_staff(values["staff_id"]) or find_digital_id_by_ms_user_id(ms_user_id)):
        return render_template(
            "index.html",
            departments=DEPARTMENTS, employment_statuses=EMPLOYMENT_STATUSES, genders=GENDERS, categories=CATEGORIES,
            error="A Digital ID already exists for this Staff ID. Use \u201cExisting Staff Activation\u201d instead.",
            form=form,
        ), 409

    token = edit_record["Token"] if edit_record else secrets.token_urlsafe(24)

    photo_filename = edit_record.get("Photo Filename") if edit_record else None
    if photo_provided:
        photo_filename, photo_error = _validate_and_save_photo(photo_file, token)
        if photo_error:
            return render_template(
                "index.html",
                departments=DEPARTMENTS, employment_statuses=EMPLOYMENT_STATUSES, genders=GENDERS, categories=CATEGORIES,
                error=photo_error,
                form=form, edit_record=edit_record,
            ), 400

    if edit_record:
        update_digital_id_record(token, {
            "First Name": values["first_name"], "Last Name": values["last_name"],
            "Email": values["email"], "Mobile Number": mobile_number,
            "Department": values["department"], "Job Title": values["job_title"],
            "Gender": values["gender"], "Employment Status": values["employment_status"],
            "Photo Filename": photo_filename,
            "Category": category or edit_record.get("Category") or "",
            "UK IT User ID": uk_it_user_id or edit_record.get("UK IT User ID") or "",
            "Local Login": local_login or edit_record.get("Local Login") or "",
            "MISIS": misis or edit_record.get("MISIS") or "",
        })
        regenerate_digital_id_qr(token)
        log_event(
            "digital_id_updated",
            session.get("ms_user_id") or edit_record["Staff ID"],
            f"Digital ID updated for {edit_record['Staff ID']} ({values['first_name']} {values['last_name']})",
        )
        session.pop("sso_prefill", None)
        session.pop("sso_photo", None)
        return redirect(url_for("success", token=token))

    track_id = generate_track_id()

    vcard_payload = build_vcard(
        first_name=values["first_name"], last_name=values["last_name"],
        job_title=values["job_title"], department=values["department"],
        email=values["email"], mobile=mobile_number,
    )
    qr = qrcode.QRCode(version=1, box_size=10, border=3)
    qr.add_data(vcard_payload)
    qr.make(fit=True)
    # Black-on-white — every colour option gives lower contrast than this
    # for camera scanning, and reliability matters more than branding here.
    qr_img = qr.make_image(fill_color="#000000", back_color="#ffffff")
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format="PNG")
    qr_bytes = qr_buf.getvalue()

    if storage.blob_configured():
        qr_value = storage.upload_qr(qr_bytes, token)
    else:
        qr_filename = f"qr_{token}.png"
        with open(os.path.join(QR_DIR, qr_filename), "wb") as f:
            f.write(qr_bytes)
        qr_value = qr_filename

    record = {
        "Token": token,
        "Track ID": track_id,
        "Staff ID": values["staff_id"],
        "First Name": values["first_name"],
        "Last Name": values["last_name"],
        "Email": values["email"],
        "Mobile Number": mobile_number,
        "Department": values["department"],
        "Job Title": values["job_title"],
        "Gender": values["gender"],
        "Employment Status": values["employment_status"],
        "Photo Filename": photo_filename,
        "QR Filename": qr_value,
        "Created At": datetime.now().isoformat(timespec="seconds"),
        "Status": "Active",
        "Microsoft User ID": ms_user_id,
        "Category": category,
        "UK IT User ID": uk_it_user_id,
        "Local Login": local_login,
        "MISIS": misis,
    }
    append_digital_id(record)
    log_event(
        "digital_id_created",
        ms_user_id or values["staff_id"],
        f"Digital ID created for {values['staff_id']} ({values['first_name']} {values['last_name']})",
    )
    session.pop("sso_prefill", None)
    session.pop("sso_photo", None)

    return redirect(url_for("success", token=token))


@app.route("/healthz")
def healthz():
    info = {
        "status": "ok",
        "db_configured": storage.db_configured(),
        "blob_configured": storage.blob_configured(),
    }
    if storage.db_configured():
        try:
            conn = storage.get_connection()
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM digital_ids")
                    info["digital_ids_row_count"] = cur.fetchone()[0]
            finally:
                conn.close()
        except Exception as exc:
            info["db_error"] = str(exc)
    return jsonify(info)


@app.route("/")
def index():
    """Landing page: Microsoft sign-in first. Staff who'd rather not use
    Microsoft (or whose account isn't in Entra ID yet) can still reach the
    manual registration/Track ID lookup page directly via /portal."""
    return render_template(
        "login.html",
        sso_configured=sso_config.is_enabled(),
        sso_error=request.args.get("sso_error") or session.pop("sso_error", None),
        flash_toast=session.pop("flash_toast", None),
    )


def _edit_prefill_profile(record):
    return {
        "first_name": record["First Name"], "last_name": record["Last Name"],
        "staff_id": record["Staff ID"], "email": record["Email"],
        "mobile_number": record.get("Mobile Number") or "",
        "department": record["Department"], "job_title": record["Job Title"],
        "gender": record.get("Gender") or "", "employment_status": record.get("Employment Status") or "",
        "category": record.get("Category") or "",
        "uk_it_user_id": record.get("UK IT User ID") or "",
        "local_login": record.get("Local Login") or "",
        "misis": record.get("MISIS") or "",
    }


@app.route("/portal")
def portal():
    edit_token = request.args.get("edit")
    edit_record = None
    sso_prefill = None
    sso_photo = None

    if edit_token:
        edit_record = find_digital_id_by_token(edit_token)
        if not edit_record or edit_record.get("Microsoft User ID") != session.get("ms_user_id"):
            abort(403)
        sso_prefill = _edit_prefill_profile(edit_record)
        if edit_record.get("Photo Filename"):
            sso_photo = url_for("serve_photo", token=edit_token)
    else:
        sso_prefill = session.pop("sso_prefill", None)
        sso_photo = sso_photo_cache.pop_photo(session.get("ms_user_id"))

    return render_template(
        "index.html",
        departments=DEPARTMENTS, employment_statuses=EMPLOYMENT_STATUSES, genders=GENDERS, categories=CATEGORIES,
        show_sso_banner=sso_config.is_enabled() and not session.get("ms_user_id") and not edit_record,
        sso_error=session.pop("sso_error", None),
        sso_prefill=sso_prefill,
        sso_photo=sso_photo,
        edit_record=edit_record,
        signed_in=bool(session.get("ms_user_id")),
    )


@app.route("/account")
def account():
    """Post-sign-in welcome hub: greets the staff member and routes them to
    either their existing Digital ID or into a prefilled registration form,
    depending on whether one already exists."""
    ms_user_id = session.get("ms_user_id")
    if not ms_user_id:
        return redirect(url_for("index"))

    profile = session.get("sso_prefill") or {}
    existing = find_digital_id_by_ms_user_id(ms_user_id)
    if not existing and profile.get("staff_id"):
        existing = find_active_digital_id_by_staff(profile["staff_id"])

    display_name = profile.get("full_name") or profile.get("first_name") or ""
    if existing and not display_name:
        display_name = f"{existing['First Name']} {existing['Last Name']}".strip()

    return render_template(
        "account.html",
        display_name=display_name or "there",
        profile=profile,
        photo=sso_photo_cache.get_photo(ms_user_id),
        existing=existing if existing and existing.get("Status") == "Active" else None,
    )


@app.route("/register", methods=["POST"])
def register():
    return _handle_submission()


@app.route("/activate", methods=["POST"])
def activate():
    return _handle_submission()


@app.route("/api/track", methods=["POST"])
def api_track():
    payload = request.get_json(silent=True) or {}
    query = (payload.get("query") or "").strip()

    if not query:
        return jsonify({"found": False, "error": "Enter a Track ID or Staff ID."}), 400

    digital = find_digital_id_by_track(query) or find_active_digital_id_by_staff(query)
    if digital:
        return jsonify({
            "found": True,
            "type": "digital_id",
            "record": digital,
            "photo_url": url_for("serve_photo", token=digital["Token"]) if digital.get("Photo Filename") else None,
            "success_url": url_for("success", token=digital["Token"]),
        })

    directory = find_directory_record_by_staff_id(query)
    if directory:
        return jsonify({"found": True, "type": "directory", "record": directory})

    return jsonify({"found": False})


@app.route("/success/<token>")
def success(token):
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)

    verify_url = f"{PORTAL_BASE_URL}/verify/{token}"
    wallet_url = wallet.build_save_url(
        staff_id=record["Staff ID"],
        full_name=f"{record['First Name']} {record['Last Name']}",
        job_title=record["Job Title"],
        department=record["Department"],
        employment_status=record["Employment Status"],
        verify_url=verify_url,
    )
    can_edit = bool(record.get("Microsoft User ID") and record.get("Microsoft User ID") == session.get("ms_user_id"))
    return render_template(
        "success.html",
        record=record, token=token, verify_url=verify_url, can_edit=can_edit,
        wallet_url=wallet_url, wallet_configured=wallet.is_configured(),
        apple_configured=wallet_apple.is_configured(),
        active_count=count_active_digital_ids(),
    )


@app.route("/verify/<token>")
def verify(token):
    record = find_digital_id_by_token(token)
    if not record:
        return render_template("verify.html", status="invalid")

    record_status = record.get("Status") or "Active"
    status = record_status.lower() if record_status in DIGITAL_ID_STATUSES else "invalid"
    return render_template("verify.html", status=status, record=record)


@app.route("/photos/<token>")
def serve_photo(token):
    """Proxies the staff photo by Digital ID token rather than serving a raw
    filename/URL directly -- the Blob store is private, so the browser can't
    fetch it itself; this route attaches the read-write token server-side."""
    record = find_digital_id_by_token(token)
    data = _load_media_bytes(record.get("Photo Filename"), PHOTO_DIR) if record else None
    if not data:
        abort(404)
    return send_file(data, mimetype="image/jpeg")


@app.route("/qrcodes/<token>")
def serve_qr(token):
    record = find_digital_id_by_token(token)
    data = _load_media_bytes(record.get("QR Filename"), QR_DIR) if record else None
    if not data:
        abort(404)
    return send_file(data, mimetype="image/png")


@app.route("/preview/<token>")
def preview_card(token):
    """Serves the front of the rendered card image used by /download,
    inline (not as an attachment) so the success page can show staff the
    real final Digital ID -- not just a CSS approximation of it -- before
    they download or add it to a wallet."""
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)

    photo_file = _load_media_bytes(record.get("Photo Filename"), PHOTO_DIR)
    qr_file = _load_media_bytes(record.get("QR Filename"), QR_DIR)
    card = render_card_front(record, photo_file, qr_file)

    buf = io.BytesIO()
    card.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png", as_attachment=False)


@app.route("/preview/<token>/back")
def preview_card_back(token):
    """Back of the card: IT-office terms, address, and the Code128 barcode
    of the Staff ID."""
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)

    card = render_card_back(record)

    buf = io.BytesIO()
    card.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png", as_attachment=False)


@app.route("/download/<token>")
def download(token):
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)

    photo_file = _load_media_bytes(record.get("Photo Filename"), PHOTO_DIR)
    qr_file = _load_media_bytes(record.get("QR Filename"), QR_DIR)
    card = render_card_front(record, photo_file, qr_file)

    buf = io.BytesIO()
    card.save(buf, format="PNG")
    buf.seek(0)
    filename = f"MDX-Digital-ID-{record['Staff ID']}.png"
    return send_file(buf, mimetype="image/png", as_attachment=True, download_name=filename)


@app.route("/download/<token>/pdf")
def download_pdf(token):
    """A 2-page PDF -- front then back -- mirroring the official Staff ID
    badge template's own front/back page structure."""
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)

    photo_file = _load_media_bytes(record.get("Photo Filename"), PHOTO_DIR)
    qr_file = _load_media_bytes(record.get("QR Filename"), QR_DIR)
    front = render_card_front(record, photo_file, qr_file).convert("RGB")
    back = render_card_back(record).convert("RGB")

    buf = io.BytesIO()
    front.save(buf, format="PDF", save_all=True, append_images=[back])
    buf.seek(0)
    filename = f"MDX-Digital-ID-{record['Staff ID']}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/api/wallet/apple/<token>")
def wallet_apple_pass(token):
    record = find_digital_id_by_token(token)
    if not record:
        abort(404)
    if not wallet_apple.is_configured():
        abort(404)

    verify_url = f"{PORTAL_BASE_URL}/verify/{token}"
    pkpass_bytes = wallet_apple.build_pkpass_bytes(
        staff_id=record["Staff ID"],
        full_name=f"{record['First Name']} {record['Last Name']}",
        job_title=record["Job Title"],
        department=record["Department"],
        employment_status=record["Employment Status"],
        verify_url=verify_url,
        status=record.get("Status", "Active"),
    )
    if not pkpass_bytes:
        abort(404)

    buf = io.BytesIO(pkpass_bytes)
    filename = f"MDX-Digital-ID-{record['Staff ID']}.pkpass"
    return send_file(buf, mimetype="application/vnd.apple.pkpass", as_attachment=True, download_name=filename)


def _lan_ip():
    """Best-effort LAN-facing IP for convenience — doesn't actually send
    any traffic, just asks the OS which local interface would be used to
    reach the internet."""
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(("8.8.8.8", 80))
        return s.getsockname()[0]
    except OSError:
        return None
    finally:
        s.close()


if __name__ == "__main__":
    lan_ip = _lan_ip()
    print(" * Local:   http://127.0.0.1:5000")
    if lan_ip:
        print(f" * Network: http://{lan_ip}:5000  (open this on a phone/device on the same Wi-Fi/LAN)")
    else:
        print(" * Network: could not detect a LAN IP — check `ipconfig` for your adapter's IPv4 address")
    print(" * If another device can't connect, allow inbound TCP port 5000 through Windows Firewall.")
    app.run(host="0.0.0.0", debug=True, port=5000)
